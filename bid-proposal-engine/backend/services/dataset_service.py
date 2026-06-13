"""Dataset Service — Single source of truth from the hackathon Excel workbook.

Loads and parses the TEKROWE workbook on startup.
All business intelligence originates from these records.
"""

import os
import openpyxl

_workbook = None
_bid_history = []
_capability_library = []

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data",
                          "Problem#1_Sample_Datasets (TEKROWE).xlsx")

# Canonical column names (row index 1 in each sheet, 0-indexed)
BID_HISTORY_COLS = [
    "bid_id", "client", "sector", "budget", "score_pct",
    "outcome", "response_time_hrs", "compliance_pct",
    "doc_pages", "gaps_found", "bid_manager", "submission_date"
]

CAPABILITY_COLS = [
    "cap_id", "domain", "project_summary", "certification",
    "year_completed", "contract_value", "duration_months", "client_type"
]


def _load_workbook():
    """Load and parse the Excel workbook into memory."""
    global _bid_history, _capability_library

    if _bid_history:  # Already loaded
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

    # --- Bid History ---
    ws_bh = wb[wb.sheetnames[0]]
    raw_bh = list(ws_bh.iter_rows(values_only=True))
    # Row 0 = title, Row 1 = headers, Row 2+ = data
    for row in raw_bh[2:]:
        if not row[0] or str(row[0]).strip() == "":
            continue
        record = {}
        for i, col in enumerate(BID_HISTORY_COLS):
            val = row[i] if i < len(row) else None
            record[col] = val
        _bid_history.append(record)

    # --- Capability Library ---
    ws_cl = wb[wb.sheetnames[1]]
    raw_cl = list(ws_cl.iter_rows(values_only=True))
    for row in raw_cl[2:]:
        if not row[0] or str(row[0]).strip() == "":
            continue
        record = {}
        for i, col in enumerate(CAPABILITY_COLS):
            val = row[i] if i < len(row) else None
            record[col] = val
        _capability_library.append(record)

    wb.close()


def get_bid_history() -> list[dict]:
    """Return all bid history records."""
    _load_workbook()
    return _bid_history


def get_capability_records() -> list[dict]:
    """Return all capability library records."""
    _load_workbook()
    return _capability_library


def get_evaluation_taxonomy() -> dict:
    """Return derived taxonomy: sectors, domains, certifications, client types."""
    _load_workbook()

    sectors = sorted(set(r["sector"] for r in _bid_history if r["sector"]))
    domains = sorted(set(r["domain"] for r in _capability_library if r["domain"]))
    certifications = sorted(set(r["certification"] for r in _capability_library
                                if r["certification"] and r["certification"] != "N/A"))
    client_types = sorted(set(r["client_type"] for r in _capability_library if r["client_type"]))

    return {
        "sectors": sectors,
        "domains": domains,
        "certifications": certifications,
        "client_types": client_types,
    }


def get_capability_text(record: dict) -> str:
    """Build a text representation of a capability record for embedding."""
    parts = [
        f"Domain: {record.get('domain', '')}",
        f"Project: {record.get('project_summary', '')}",
        f"Certification: {record.get('certification', 'N/A')}",
        f"Client Type: {record.get('client_type', '')}",
        f"Contract Value: {record.get('contract_value', '')}",
        f"Year: {record.get('year_completed', '')}",
    ]
    return " | ".join(p for p in parts if p)
