"""Migration script — imports Excel workbook data into PostgreSQL/SQLite database.

Run once to populate the database with bid history and capability library data.
Also generates embeddings for capability records using sentence-transformers.

Usage:
    cd backend
    python -m db.migrate_excel
"""

import os
import sys
import json
import logging

# Add backend to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from sqlalchemy.orm import Session
from config.database import engine, SessionLocal, init_db, enable_pgvector, IS_POSTGRES
from models.bid_history import BidHistory
from models.capability_library import CapabilityLibrary

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data",
                          "Problem#1_Sample_Datasets (TEKROWE).xlsx")

BID_HISTORY_COLS = [
    "bid_id", "client", "sector", "budget", "score_pct",
    "outcome", "response_time_hrs", "compliance_pct",
    "doc_pages", "gaps_found", "bid_manager", "submission_date"
]

CAPABILITY_COLS = [
    "cap_id", "domain", "project_summary", "certification",
    "year_completed", "contract_value", "duration_months", "client_type"
]


def get_capability_text(record: dict) -> str:
    """Build text representation for embedding."""
    parts = [
        f"Domain: {record.get('domain', '')}",
        f"Project: {record.get('project_summary', '')}",
        f"Certification: {record.get('certification', 'N/A')}",
        f"Client Type: {record.get('client_type', '')}",
        f"Contract Value: {record.get('contract_value', '')}",
        f"Year: {record.get('year_completed', '')}",
    ]
    return " | ".join(p for p in parts if p)


def load_excel_data():
    """Load data from Excel workbook."""
    logger.info(f"Loading Excel from: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

    # Bid History (Sheet 0)
    ws_bh = wb[wb.sheetnames[0]]
    raw_bh = list(ws_bh.iter_rows(values_only=True))
    bid_history = []
    for row in raw_bh[2:]:
        if not row[0] or str(row[0]).strip() == "":
            continue
        # Skip header-like rows
        if str(row[0]).strip().lower() in ("bid_id", "bid id", "id", ""):
            continue
        record = {}
        for i, col in enumerate(BID_HISTORY_COLS):
            val = row[i] if i < len(row) else None
            record[col] = val
        bid_history.append(record)

    # Capability Library (Sheet 1)
    ws_cl = wb[wb.sheetnames[1]]
    raw_cl = list(ws_cl.iter_rows(values_only=True))
    capability_library = []
    for row in raw_cl[2:]:
        if not row[0] or str(row[0]).strip() == "":
            continue
        # Skip header-like rows
        if str(row[0]).strip().lower() in ("cap_id", "cap id", "id", ""):
            continue
        record = {}
        for i, col in enumerate(CAPABILITY_COLS):
            val = row[i] if i < len(row) else None
            record[col] = val
        capability_library.append(record)

    wb.close()
    logger.info(f"Loaded {len(bid_history)} bid records, {len(capability_library)} capability records")
    return bid_history, capability_library


def generate_embeddings(capabilities: list[dict]) -> dict:
    """Generate embeddings for capability records using sentence-transformers."""
    logger.info("Generating embeddings with sentence-transformers...")
    try:
        from sentence_transformers import SentenceTransformer

        model = SentenceTransformer("all-MiniLM-L6-v2")
        embeddings = {}

        for cap in capabilities:
            text = get_capability_text(cap)
            embedding = model.encode(text).tolist()
            embeddings[cap["cap_id"]] = (embedding, text)

        logger.info(f"Generated {len(embeddings)} embeddings (384 dimensions)")
        return embeddings
    except Exception as e:
        logger.warning(f"Embedding generation failed: {e}")
        logger.info("Continuing without embeddings (can be generated later)")
        # Return empty embeddings - store text only
        return {cap["cap_id"]: (None, get_capability_text(cap)) for cap in capabilities}


def migrate():
    """Run the full migration."""
    logger.info("Starting migration...")

    # Initialize database tables
    init_db()
    logger.info("Database tables created")

    # Enable pgvector if PostgreSQL
    if enable_pgvector():
        logger.info("pgvector extension enabled")
    else:
        logger.info("Using SQLite (pgvector not available)")

    # Load Excel data
    bid_history, capability_library = load_excel_data()

    # Generate embeddings
    embeddings = generate_embeddings(capability_library)

    # Insert data
    db = SessionLocal()
    try:
        # Check if data already exists
        existing_bids = db.query(BidHistory).count()
        existing_caps = db.query(CapabilityLibrary).count()

        if existing_bids > 0 or existing_caps > 0:
            logger.info(f"Database already has {existing_bids} bids, {existing_caps} capabilities")
            logger.info("Skipping migration (delete the database file to re-migrate)")
            return

        # Insert bid history
        for record in bid_history:
            # Parse budget string (e.g., "PKR 22M" -> store as string, not numeric)
            budget_val = record.get("budget")
            if budget_val is not None:
                budget_val = str(budget_val)

            bid = BidHistory(
                bid_id=str(record["bid_id"]),
                client=str(record.get("client", "")),
                sector=str(record.get("sector", "")),
                budget=budget_val,
                score_pct=record.get("score_pct"),
                outcome=str(record.get("outcome", "")),
                response_time_hrs=record.get("response_time_hrs"),
                compliance_pct=record.get("compliance_pct"),
                doc_pages=record.get("doc_pages"),
                gaps_found=str(record.get("gaps_found", "")),
                bid_manager=str(record.get("bid_manager", "")),
                submission_date=str(record.get("submission_date", "")),
            )
            db.add(bid)

        # Insert capability library with embeddings
        for record in capability_library:
            cap_id = str(record["cap_id"])
            embedding_data = embeddings.get(cap_id)

            # For SQLite, store embedding as JSON string
            embedding_val = None
            if embedding_data:
                if IS_POSTGRES:
                    embedding_val = embedding_data[0]  # pgvector accepts list directly
                else:
                    embedding_val = json.dumps(embedding_data[0])  # JSON string for SQLite

            cap = CapabilityLibrary(
                cap_id=cap_id,
                domain=str(record.get("domain", "")),
                project_summary=str(record.get("project_summary", "")),
                certification=str(record.get("certification", "N/A")),
                year_completed=str(record.get("year_completed", "")),
                contract_value=str(record.get("contract_value", "")),
                duration_months=record.get("duration_months"),
                client_type=str(record.get("client_type", "")),
                embedding=embedding_val,
                embedding_text=embedding_data[1] if embedding_data else None,
            )
            db.add(cap)

        db.commit()
        logger.info(f"Inserted {len(bid_history)} bid records and {len(capability_library)} capability records")

    except Exception as e:
        db.rollback()
        logger.error(f"Migration failed: {e}")
        raise
    finally:
        db.close()

    logger.info("Migration complete!")


if __name__ == "__main__":
    migrate()
